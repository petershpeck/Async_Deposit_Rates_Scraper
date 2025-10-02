import pandas as pd
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import os

OUTPUT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'output'))
os.makedirs(OUTPUT_DIR, exist_ok=True)

def save_all_to_xlsx(all_products, out_path=None):
    out_path = out_path or os.path.join(OUTPUT_DIR, 'Deposit_Rate_Data.xlsx')

    # === Формируем новые данные в DataFrame ===
    all_rows = []
    for bank, products in all_products.items():
        if not products:
            continue
        for prod in products:
            if 'bank' not in prod or not prod['bank']:
                prod['bank'] = bank
            all_rows.append(prod)

    df = pd.DataFrame(all_rows)

    # Гарантируем наличие всех колонок
    cols = [
        'bank', 'nkb', 'full_name', 'group_1', 'product',
        'date', 'day', 'month', 'year', 'week',
        'currency', 'term', 'rate', 'source_url'
    ]
    for c in cols:
        if c not in df.columns:
            df[c] = ''

    df = df[cols]

    # Добавляем текущую дату
    df['date'] = pd.to_datetime(datetime.now().date())

    # day, month, year, week
    df['day'] = df['date'].dt.day
    df['month'] = df['date'].dt.month
    df['year'] = df['date'].dt.year
    df['week'] = df['date'].dt.isocalendar().week

    # Добавляем букву "m" к term
    df['term'] = df['term'].astype(str) + 'm'

    # Конвертируем rate в float
    df['rate'] = pd.to_numeric(df['rate'], errors='coerce')

    sheet_name = "Select Rates"

    # === Проверяем, существует ли файл и лист ===
    file_exists = os.path.exists(out_path)
    if file_exists:
        workbook = load_workbook(out_path)

        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            startrow = sheet.max_row # следующая пустая строка
            # Убираем пустую строку
            if startrow > 0:
                startrow -= 1
        else:
            startrow = 0
    else:
        startrow = 0

    # === Записываем данные в Excel ===
    with pd.ExcelWriter(
        out_path,
        engine='openpyxl',
        mode='a' if file_exists else 'w',
        if_sheet_exists='overlay' if file_exists else None
    ) as writer:
        df.to_excel(
            writer,
            sheet_name=sheet_name,
            index=False,
            startrow=startrow,
            header=(startrow == 0)
        )

        worksheet = writer.sheets[sheet_name]

        # Форматирование колонки date
        date_col_idx = df.columns.get_loc('date') + 1
        date_col_letter = get_column_letter(date_col_idx)
        for row in worksheet[f"{date_col_letter}{startrow+1}:{date_col_letter}{startrow+len(df)+1}"]:
            for cell in row:
                cell.number_format = "DD.MM.YYYY"

        # Форматирование колонки rate
        rate_col_idx = df.columns.get_loc('rate') + 1
        rate_col_letter = get_column_letter(rate_col_idx)
        for row in worksheet[f"{rate_col_letter}{startrow+1}:{rate_col_letter}{startrow+len(df)+1}"]:
            for cell in row:
                cell.number_format = "0.00"

    print(f"Saved Excel -> {out_path}")
    return out_path
